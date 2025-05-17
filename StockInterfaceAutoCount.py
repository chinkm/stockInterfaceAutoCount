#import sqlalchemy as sa
import pyodbc
import pandas as pd
import numpy as np
from datetime import date
import datetime
import calendar
import sys
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font
from tkinter import messagebox
import warnings

#pd.set_option('display.max_columns', None)
#pd.set_option('display.max_rows', None)

warnings.filterwarnings("ignore")

class StockAutoCountInterface:
    def __init__(self, connection_string, companyName):
        self.connection_string=connection_string
        self.companyName=companyName

    # sqlalchemy library was not used and replaced by pyodbc so the connectionString function was unnecessary 
    #def connectionString(self):
        #connection_url=sa.engine.URL.create("access+pyodbc", query={"odbc_connect": self.connection_string})

        #self.engine=sa.create_engine(connection_url, echo=False)

    def sqlQryDB(self):

        try:
            #conn=self.engine.connect()
            conn=pyodbc.connect(self.connection_string)

            sql="SELECT * FROM [StockListT]"
            self.result=pd.read_sql(sql, conn)
            
            sql1="SELECT IIf([CompanyShort]='SE','Sxxxxi',IIf([CompanyShort]='GM','Gxxxxxxxxxxxa',IIf([CompanyShort]='ES','Exxxxxxxxxxxxxs',IIf([CompanyShort]='PE','Pxxxxxxxxxxe',IIf([CompanyShort]='FS','FSM'))))) \
                    AS Company, StockSQ.ID, StockSQ.Stock, StockSQ.Quantity, StockSQ.Unit, StockSQ.UID, StockSQ.Price1 AS Price, StockSQ.Quantity1 AS Quantity1, Price*Quantity1 AS TotalPrice, Mid([UID],1,2) AS CompanyShort, \
                    CDate(Mid([UID],5,2) & ',' & Mid([UID],7,2) & ',' & Mid([UID],9,2)) AS [Date], StockSQ.[Sub-Company] FROM StockSQ WHERE Company LIKE '"+companyName+"'"+" ORDER BY UID"
            self.result1=pd.read_sql(sql1, conn)
            
            
            sql2="SELECT IIf([CompanyShort]='SE','Sxxxxi',IIf([CompanyShort]='GM','Gxxxxxxxxxxxxa',IIf([CompanyShort]='ES','Exxxxxxxxxxxxxs',IIf([CompanyShort]='PE','Pxxxxxxxxxe',IIf([CompanyShort]='FS','FSM'))))) AS Company, \
                  StockSQ.ID, StockSQ.Stock, StockSQ.Quantity, StockSQ.Unit, StockSQ.UID, StockSQ.Price2 AS Price, StockSQ.Quantity2 AS Quantity1, Price*Quantity1 AS TotalPrice, Mid([UID],1,2) AS CompanyShort, CDate(Mid([UID],5,2) \
                  & ',' & Mid([UID],7,2) & ',' & Mid([UID],9,2)) AS [Date], StockST.[Sub-Company] FROM StockSQ WHERE StockSQ.Price2<>0 AND StockSQ.Quantity2<>0 AND Company LIKE '"+companyName+"'"+" ORDER BY UID"
            self.result2=pd.read_sql(sql2, conn)

            
            sql3="SELECT IIf([CompanyShort]='SE','Sxxxxi',IIf([CompanyShort]='GM','Gxxxxxxxxxxxxa',IIf([CompanyShort]='ES','Exxxxxxxxxxxxxxs',IIf([CompanyShort]='PE','Pxxxxxxxxxxe',IIf([CompanyShort]='FS','FSM'))))) AS Company, \
                  StockSQ.ID, StockSQ.Stock, StockSQ.Quantity, StockSQ.Unit, StockSQ.UID, StockSQ.Price3 AS Price, StockSQ.Quantity3 AS Quantity1, Price*Quantity1 AS TotalPrice, Mid([UID],1,2) AS CompanyShort, CDate(Mid([UID],5,2) \
                  & ',' & Mid([UID],7,2) & ',' & Mid([UID],9,2)) AS [Date], StockST.[Sub-Company] FROM StockSQ WHERE StockSQ.Price3<>0 AND StockSQ.Quantity3<>0 AND Company LIKE '"+companyName+"'"+" ORDER BY UID"
            self.result3=pd.read_sql(sql3, conn)


            frames=[self.result1, self.result2, self.result3]

            self.df_concat=pd.concat(frames, axis=0, ignore_index=True)
            
            
            sql5="SELECT * FROM [BlockListT] WHERE Company LIKE '"+companyName+"'"
            self.result5=pd.read_sql(sql5, conn)

            sql6=r"SELECT MasterQ.[Combined_ID], MasterQ.[Company], MasterQ.Day, MasterQ.Block, MasterQ.Task, MasterQ.Job, TaskListT.[Account_Code], StockListT.[Stock Type] FROM (([MasterQ] INNER JOIN StockSQ ON MasterQ.[Combined_ID]=StockSQ.UID) INNER JOIN TaskListT ON MasterQ.[Task]=TaskListT.[Task]) INNER JOIN StockListT ON StockSQ.Stock=StockListT.[Stock Name] WHERE Company LIKE '"+companyName+"'"+" ORDER BY Combined_ID"
            self.result6=pd.read_sql(sql6, conn)

        except Exception as e:
            messagebox.showinfo(e.__class__, e.orig)

        finally:
            conn.close()


    
    def generateAccountCode(self):

        if self.result.empty:
            messagebox.showinfo("askquestion", self.companyName+" has no data in the database, Are you sure you have chosen the right company?")
            sys.exit(0)
                
        status={"Mature":"600", "Immature": "220", "Others": "710"}
        # match Combined_ID from result6 to UID from df_concat
        for index, row in self.df_concat.iterrows():
            # generate AccountName & Block Column in df_concat dataset
            if self.result6["Combined_ID"].str.contains(row["UID"]).any():
                self.df_concat.loc[index, "AccountName"]=self.result6[self.result6["Combined_ID"].str.strip()==row["UID"]]["Task"].values[0]
                self.df_concat.loc[index, "Block"]=self.result6[self.result6["Combined_ID"].str.strip()==row["UID"]]["Block"].values[0]
                self.df_concat.loc[index, "AccountCode"]=self.result6[self.result6["Combined_ID"].str.strip()==row['UID']]["Account_Code"].values[0]
                self.df_concat.loc[index, "Stock Type"]=self.result6[self.result6["Combined_ID"].str.strip()==row["UID"]]["Stock Type"].values[0]

        # generate the status code e.g Mature/Immature/Others
        for index, row in self.df_concat.iterrows():
            if self.result5["Block"].eq(row["Block"]).any():
                self.df_concat.loc[index, "Status"]=self.result5[self.result5["Block"]==row["Block"]]["Status"].values[0]
            elif row["Block"].strip()=="OTHERS":
                self.df_concat.loc[index, "Status"]="Others"
                                       
        # generate Account Code
        for index, row in self.df_concat.iterrows():
            if row["Status"]=="Mature":
                self.df_concat.loc[index, "Account_Code"]=status["Mature"]+"-"+row["AccountCode"]
            elif row["Status"]=="Immature":
                self.df_concat.loc[index, "Account_Code"]=status["Immature"]+"-"+row["AccountCode"]
            elif row["Status"]=="Others":
                self.df_concat.loc[index, "Account_Code"]=status["Others"]+"-"+row["AccountCode"]

        for index, row in self.df_concat.iterrows():
            if row["Block"]=="OTHERS" and row["AccountCode"]=="U001":
                self.df_concat.loc[index, "Account_Code"]="700-U001"
            elif row["Block"]=="OTHERS":
                self.df_concat.loc[index, "Account_Code"]="700-U001"
            elif (bool(re.search("^P3.*",row["AccountCode"]))):
                self.df_concat.loc[index, "Account_Code"]="220-"+row["AccountCode"]

        # make uppercase AccountName column description
        self.df_concat["AccountName"]=self.df_concat["AccountName"].str.upper()

        # generate Month Column in df_concat
        self.df_concat["Month"]=self.df_concat["Date"].apply(lambda x: x.strftime('%B %Y'))
        
          

    def generateJVInterface(self):
        fontStyle=Font(name="Tahoma", size=8)
        groupbyResult=self.df_concat.groupby(["AccountName", "Block", "Account_Code", "Status", "Stock", "Unit", "Price", "Stock Type"], dropna=False).aggregate({"TotalPrice":"sum", "Quantity1":"sum"})
        updateResult=groupbyResult.reset_index().replace({np.nan:''})
        wb=load_workbook(filename="C:\\Users\\User\\Desktop\\WorkingFolder\\Project_ReadMasterSheet\\Import Journal Entry-Stock.xlsx")
                
            
        sheet2=wb["Sheet2"]
        sheet1=wb.copy_worksheet(sheet2)
        rows=4
        
        for index, row in updateResult.iterrows():
            if row["Stock Type"]=="Chemical":
                sheet1["F"+str(rows)]=row['AccountName']+" - " +row["Stock"]+" "+"{:.3f}".format(row['Quantity1'])+" "+row['Unit']+" @ RM"+str(row['Price'])
                sheet1["F"+str(rows)].font=fontStyle
            else:
                sheet1["F"+str(rows)]="MANURING "+row['AccountName']+" - "+"{:.3f}".format(row['Quantity1'])+" "+row['Unit']+" @ RM"+str(row['Price'])
                sheet1["F"+str(rows)].font=fontStyle
            sheet1["K"+str(rows)]=row['Account_Code']
            sheet1["K"+str(rows)].font=fontStyle
            sheet1["M"+str(rows)]=row['Block']
            sheet1["M"+str(rows)].font=fontStyle
            sheet1["N"+str(rows)]="MATERIAL"
            sheet1["N"+str(rows)].font=fontStyle
            sheet1["R"+str(rows)]="{:.2f}".format(row['Quantity1'])+" "+row['Unit']
            sheet1["R"+str(rows)].font=fontStyle
            sheet1["Y"+str(rows)]=float("{:.2f}".format(row['TotalPrice']))
            sheet1["Y"+str(rows)].number_format="0.00"
            sheet1["Y"+str(rows)].font=fontStyle
            rows=rows+1

        
        sheet1["P4"]="BEING FERTILIZER & CHEMICAL ISSUED FOR THE MONTH " +self.df_concat["Month"].values[0].upper()
        sheet1["P4"].font=fontStyle
        sheet1["C4"]=str(calendar.monthrange(int(self.df_concat['Date'].dt.strftime('%Y').values[0]), int(self.df_concat['Date'].dt.strftime('%m').values[0]))[1])+"/"+str(self.df_concat['Date'].dt.strftime('%m').values[0])+"/"+str(self.df_concat['Date'].dt.strftime('%Y').values[0])
        sheet1["C4"].font=fontStyle                                     
        
        # Segregating out Fertilizer and Chemical Stock listed in the JV Interface Report
        groupbyResult=self.df_concat.groupby(["Stock Type", "Stock", "Unit", "Price"], dropna=False).aggregate({"TotalPrice":"sum", "Quantity1":"sum"})
        updateResult=groupbyResult.reset_index().replace({np.nan:''})
        
        for index, row in updateResult.iterrows():
            if row["Stock Type"]=="Chemical":
                sheet1["Z"+str(rows)]=round(row["TotalPrice"],2)
                sheet1["Z"+str(rows)].number_format="0.00"
                sheet1["Z"+str(rows)].font=fontStyle
                sheet1["K"+str(rows)]="300-C001"
                sheet1["K"+str(rows)].font=fontStyle
                sheet1["F"+str(rows)]=row["Stock"].upper()+" "+str(round(row["Quantity1"],2))+" "+row["Unit"][0]+"@ RM"+str(round(row["Price"],3))+" ISSUED FOR "+str(self.df_concat['Date'].dt.strftime('%m').values[0])+"/"+str(self.df_concat['Date'].dt.strftime('%Y').values[0])
                sheet1["F"+str(rows)].font=fontStyle
                rows=rows+1
            elif row["Stock Type"]=="Fertiliser":
                sheet1["Z"+str(rows)]=round(row["TotalPrice"],2)
                sheet1["Z"+str(rows)].number_format="0.00"
                sheet1["Z"+str(rows)].font=fontStyle
                sheet1["K"+str(rows)]="300-F001"
                sheet1["K"+str(rows)].font=fontStyle
                sheet1["F"+str(rows)]=row["Stock"].upper()+" "+str(round(row["Quantity1"],2))+" "+row["Unit"][0]+"@ RM"+str(round(row["Price"],3))+" ISSUED FOR "+str(self.df_concat['Date'].dt.strftime('%m').values[0])+"/"+str(self.df_concat['Date'].dt.strftime('%Y').values[0])
                sheet1["F"+str(rows)].font=fontStyle
                rows=rows+1
           
        sheet1["Y"+str(rows+8)]="=SUM(Y4:Y"+str(rows+7)+")"
        sheet1["Z"+str(rows+8)]="=SUM(Z4:Z"+str(rows+7)+")"

        sheet1.title=self.companyName+"_"+self.df_concat["Month"].values[0]
        wb.active=len(wb.sheetnames)-1
        try:
            wb.save("Import Journal Entry-Stock.xlsx")
            os.startfile("Import Journal Entry-Stock.xlsx")
        except IOError:
            messagebox.showinfo("File Open Error", "Someone has opened the Excel Import Journal Entry file. Close the file and try again")
        
            
        
       
        


companyName=sys.argv[1].strip('\"')
fullPath=sys.argv[2].strip('\"')

'''
connection_string=(
    r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};"
    r"DBQ="+fullPath+';"'
    r"ExtendedAnsiSQL=1;")
'''

connection_string=(
    r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};"
    r"DBQ="+fullPath+';"'
    r"Provider=MSDASQL")


obj=StockAutoCountInterface(connection_string, companyName)

#obj.connectionString()

obj.sqlQryDB()

obj.generateAccountCode()

obj.generateJVInterface()





            



            
            
        
        
